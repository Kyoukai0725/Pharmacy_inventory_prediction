"""
Stream-read hospital dispensing Excel exports and aggregate to drug x Monday week.

Column indices (0-based): C=stock, D=drugid, G=quantity, J=billing time, K=movement time.
Excludes stock DYG; weeks anchored on Monday; prefer column K, else J.
"""

from __future__ import annotations

from collections import defaultdict
from pathlib import Path
import numpy as np
import pandas as pd

COL_STOCK = 2
COL_DRUGID = 3
COL_QTY = 6
COL_BILL_TIME = 9
COL_MOVE_TIME = 10


def aggregate_excel_to_weekly(path: Path) -> pd.DataFrame:
    """Stream all sheets; return drugid, Monday week, and weekly_qty."""
    try:
        from openpyxl import load_workbook
    except ImportError as e:  # pragma: no cover
        raise RuntimeError("openpyxl required: pip install openpyxl") from e

    path = Path(path)
    print(f"Streaming Excel: {path.name}", flush=True)
    agg: dict[tuple[str, pd.Timestamp], float] = defaultdict(float)
    wb = load_workbook(path, read_only=True, data_only=True)
    total_rows = 0
    try:
        snames = wb.sheetnames
        for si, ws in enumerate(wb.worksheets):
            print(f"  sheet {si + 1}/{len(snames)}: {ws.title}", flush=True)
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if i == 0:
                    continue
                if not row or len(row) <= COL_MOVE_TIME:
                    continue
                stock = row[COL_STOCK]
                if stock == "DYG":
                    continue
                drugid = row[COL_DRUGID]
                if drugid is None or (isinstance(drugid, float) and np.isnan(drugid)):
                    continue
                drugid = str(drugid).strip()
                qty = row[COL_QTY]
                try:
                    q = float(qty)
                except (TypeError, ValueError):
                    continue
                dt = row[COL_MOVE_TIME] or row[COL_BILL_TIME]
                if dt is None:
                    continue
                ts = pd.Timestamp(dt)
                if pd.isna(ts):
                    continue
                week = (ts - pd.to_timedelta(ts.weekday(), unit="D")).normalize()
                agg[(drugid, week)] += q
                total_rows += 1
    finally:
        wb.close()
    if not agg:
        raise ValueError("Empty Excel aggregation; check column indices or stock filter.")
    rows = [{"drugid": k[0], "week": k[1], "weekly_qty": v} for k, v in agg.items()]
    out = pd.DataFrame(rows).sort_values(["drugid", "week"]).reset_index(drop=True)
    return out
